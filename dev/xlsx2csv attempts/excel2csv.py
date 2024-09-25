import getopt, openpyxl, csv, os.path
from openpyxl import load_workbook

import sys
sys.path.append('.')
from tools_v2 import myprint, Timer, interrupt, countLinesInCSVFile

#================================================= convert
def convert(input, worksheet, output, separator):
    wb = load_workbook(filename=input)
    #sh = wb.active
    sh = wb[worksheet]
    interrupt(sh)
    with open(output, 'w', newline="") as file_handle:
        csv_writer = csv.writer(file_handle, delimiter=separator)
        i = 0
        for row in sh.iter_rows(): # generator; was sh.rows
            print(str(i), end="|")
            csv_writer.writerow([cell.value for cell in row])

            
#================================================= usage
def usage():
    print("Usage:")
    print('> python excel2csv -i [input.xls] -w [worksheet] -o [output.csv] -s ","')
    print("i: input file (mandatory)")
    print("w: worksheet name (mandatory)")
    print("o: outputfile (mandatory)")
    print("s: separator (optional)")
    exit()


#================================================= main
def main():
    try:
        opts, args = getopt.getopt(sys.argv[1:], "i:w:o:h",
                                   ["input=", "worksheet=",
                                    "output=", "help"])
    except getopt.GetoptError:
        usage()
    if len(opts) == 0:
        usage()
    input = ""
    worksheet = ""
    output = ""
    separator = ","
    for o, a in opts:
        if o in ("-h", "--help"):
            usage()
        elif o in ("-i", "--input"):
            input = a
            print("Input file: " + a)
        elif o in ("-w", "--worksheet"):
            worksheet = a
            print("Worksheet: " + a)
        elif o in ("-o", "--output"):
            output = a
            print("Output file: " + a)
        elif o in ("-s", "--separator"):
            separator = a
            print('separator: "' + a + '"')
        else:
            print("Error: unrecognized option: '" + o + "'")
    if input == "" or worksheet == "" or output == "":
        print("Missing parameter. Exiting...")
        usage()
    if not os.path.isfile(input):
        print("File " + input + "Not found. Exiting...")
        exit()
    # main loop
    print("Extracting...")
    convert(input, worksheet, output, separator)
    print("Done")

if __name__ == '__main__':
    main()
